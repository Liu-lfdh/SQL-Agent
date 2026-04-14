import os
from langchain_core.tools import tool


@tool
def excel_reader(file_path: str) -> str:
    """读取Excel(.xlsx)文件，返回列名、前5行样例数据、总行数和每列数据类型推断
    Args:
        file_path: Excel文件的绝对路径或相对于项目根目录的路径
    """
    print(f"[excel_reader输入] 文件: {file_path}")

    if not os.path.isabs(file_path):
        file_path = os.path.abspath(file_path)

    if not os.path.exists(file_path):
        result = f"错误: 文件不存在: {file_path}"
        print(f"[excel_reader输出] {result}")
        return result

    if not file_path.endswith(".xlsx"):
        result = "错误: 仅支持 .xlsx 格式的Excel文件"
        print(f"[excel_reader输出] {result}")
        return result

    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        if ws is None:
            result = "错误: 无法读取Excel工作表"
            print(f"[excel_reader输出] {result}")
            wb.close()
            return result

        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            result = "错误: Excel文件为空"
            print(f"[excel_reader输出] {result}")
            wb.close()
            return result

        columns = [str(h) for h in rows[0]]
        data_rows = rows[1:]
        total_rows = len(data_rows)

        if total_rows == 0:
            result = "错误: Excel文件只有表头，没有数据行"
            print(f"[excel_reader输出] {result}")
            wb.close()
            return result

        # 推断每列的数据类型
        dtypes = {}
        for col_idx in range(len(columns)):
            vals = [row[col_idx] for row in data_rows if row[col_idx] is not None]
            if not vals:
                dtypes[columns[col_idx]] = "unknown"
                continue
            type_counts = {}
            for v in vals:
                t = type(v).__name__
                type_counts[t] = type_counts.get(t, 0) + 1
            dominant = max(type_counts, key=type_counts.get)
            dtypes[columns[col_idx]] = dominant

        # 构建输出：列名 + 前5行样例 + 总行数 + 类型推断
        sample_count = min(5, total_rows)
        lines = []
        lines.append(f"总行数: {total_rows}")
        lines.append(f"列名: {columns}")
        lines.append(f"列数据类型: {dtypes}")
        lines.append(f"\n前{sample_count}行样例:")
        for i in range(sample_count):
            lines.append(f"  行{i+1}: {list(data_rows[i])}")

        result = "\n".join(lines)
        print(f"[excel_reader输出] 成功读取Excel，{total_rows}行 x {len(columns)}列")
        wb.close()
        return result
    except Exception as e:
        result = f"读取Excel失败: {str(e)}"
        print(f"[excel_reader输出] {result}")
        return result
