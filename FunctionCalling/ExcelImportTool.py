import os
import time
from datetime import datetime
from langchain_core.tools import tool
from openpyxl import load_workbook

from Database_Data.Database import get_db


@tool
def excel_import(file_path: str, table_name: str, column_mapping: dict) -> str:
    """将Excel(.xlsx)文件数据批量导入到指定的数据库表中，使用参数化查询和分批插入。

    Args:
        file_path: Excel文件的绝对路径或相对于项目根目录的路径
        table_name: 目标数据库表名（如 coupon_info）
        column_mapping: 已确认的列映射字典，格式为 {"Excel列名": "数据库字段名", ...}
    """
    print(f"[excel_import输入] 文件: {file_path}, 表: {table_name}")

    # --- 安全校验（同 excel_reader）---
    target = os.path.normpath(file_path)
    abs_target = os.path.abspath(target)
    abs_root = os.path.abspath(".")

    if not abs_target.startswith(abs_root + os.sep) and abs_target != abs_root:
        result = f"错误: 不允许读取项目根目录以外的文件: {file_path}"
        print(f"[excel_import输出] {result}")
        return result

    if not os.path.exists(target):
        result = f"错误: 文件不存在: {file_path}"
        print(f"[excel_import输出] {result}")
        return result

    if not os.path.isfile(target):
        result = f"错误: 路径不是文件: {file_path}"
        print(f"[excel_import输出] {result}")
        return result

    if not target.lower().endswith(".xlsx"):
        result = "错误: 仅支持 .xlsx 格式的Excel文件"
        print(f"[excel_import输出] {result}")
        return result

    if not column_mapping:
        result = "错误: 列映射为空"
        print(f"[excel_import输出] {result}")
        return result

    # --- 验证表结构和映射字段 ---
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(f"DESC `{table_name}`")
        db_columns = {row[0] for row in cursor.fetchall()}
    except Exception as e:
        result = f"错误: 无法获取表结构: {e}"
        print(f"[excel_import输出] {result}")
        return result

    for db_col in column_mapping.values():
        if db_col not in db_columns:
            result = f"错误: 数据库表 '{table_name}' 中不存在字段 '{db_col}'"
            print(f"[excel_import输出] {result}")
            return result

    # --- 构建参数化 SQL ---
    db_fields = list(column_mapping.values())
    placeholders = ", ".join(["%s"] * len(db_fields))
    columns_sql = ", ".join([f"`{c}`" for c in db_fields])
    sql = f"INSERT INTO `{table_name}` ({columns_sql}) VALUES ({placeholders})"

    # --- 迭代读取 Excel + 分批插入 ---
    BATCH_SIZE = 1000
    wb = None
    start_time = time.time()
    total_imported = 0
    total_skipped = 0
    errors = []

    try:
        wb = load_workbook(target, read_only=True, data_only=True)
        ws = wb.active

        if ws is None:
            result = "错误: 无法读取Excel工作表"
            print(f"[excel_import输出] {result}")
            return result

        row_iter = ws.iter_rows(values_only=True)

        # 读取表头
        try:
            header_row = next(row_iter)
        except StopIteration:
            result = "错误: Excel文件为空"
            print(f"[excel_import输出] {result}")
            return result

        excel_columns = [str(h) for h in header_row]

        # 建立 Excel列索引 → 数据库字段索引 的映射
        col_indices = []
        for excel_col in column_mapping.keys():
            if excel_col not in excel_columns:
                result = f"错误: Excel文件中找不到列 '{excel_col}'"
                print(f"[excel_import输出] {result}")
                return result
            col_indices.append(excel_columns.index(excel_col))

        # 逐批读取并插入
        batch = []
        row_num = 1  # 数据行号（不含表头）

        for row in row_iter:
            row_num += 1
            try:
                values = []
                for idx in col_indices:
                    val = row[idx] if idx < len(row) else None
                    values.append(_convert_value(val))
                batch.append(tuple(values))
            except Exception as e:
                total_skipped += 1
                errors.append(f"行{row_num}: {e}")
                if len(errors) > 100:
                    errors.append("... (省略更多错误)")
                    break
                continue

            if len(batch) >= BATCH_SIZE:
                try:
                    cursor.executemany(sql, batch)
                    conn.commit()
                    total_imported += len(batch)
                except Exception as e:
                    conn.rollback()
                    total_skipped += len(batch)
                    errors.append(f"批次插入失败(行{row_num - len(batch)}~{row_num}): {e}")
                batch = []

        # 插入剩余数据
        if batch:
            try:
                cursor.executemany(sql, batch)
                conn.commit()
                total_imported += len(batch)
            except Exception as e:
                conn.rollback()
                total_skipped += len(batch)
                errors.append(f"最后一批插入失败(行{row_num - len(batch)}~{row_num}): {e}")

    except Exception as e:
        result = f"错误: 导入过程中发生异常: {e}"
        print(f"[excel_import输出] {result}")
        return result
    finally:
        if wb is not None:
            wb.close()

    elapsed = time.time() - start_time

    # --- 构建返回结果 ---
    lines = []
    lines.append(f"导入完成!")
    lines.append(f"  成功: {total_imported} 行")
    if total_skipped > 0:
        lines.append(f"  跳过: {total_skipped} 行")
    lines.append(f"  耗时: {elapsed:.2f} 秒")

    if errors:
        lines.append(f"\n错误详情(最多显示100条):")
        for err in errors[:20]:
            lines.append(f"  - {err}")

    result = "\n".join(lines)
    print(f"[excel_import输出] {total_imported} 行导入成功, {total_skipped} 行跳过, 耗时 {elapsed:.2f}s")
    return result


def _convert_value(val):
    """将 Excel 单元格的值转换为适合 pymysql 参数化查询的类型。"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(val, (int, float)):
        # 处理 Excel 中的日期序列号（如 45000.0 表示某个日期）
        if isinstance(val, float) and val > 40000:
            try:
                dt = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(val) - 2)
                return dt.strftime('%Y-%m-%d %H:%M:%S')
            except (ValueError, OverflowError):
                pass
        return val
    s = str(val).strip()
    if s == "":
        return None
    return s
